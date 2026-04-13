#!/usr/bin/env python3
"""Import data from Excel file into database"""

import sys
import openpyxl
from pathlib import Path

# Add parent to path
sys.path.insert(0, str(Path(__file__).parent))

from app import create_app, db
from app.models import User, Category, InventoryItem

def import_from_excel(excel_path):
    """Import inventory data from Excel file"""
    
    app = create_app()
    
    with app.app_context():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['چارتەکان']  # The data sheet
        
        # Create categories and items
        categories_data = {}
        
        for row_idx in range(2, ws.max_row + 1):
            col_a = ws.cell(row=row_idx, column=1).value  # Category
            col_b = ws.cell(row=row_idx, column=2).value  # System Name (or Expected)
            col_c = ws.cell(row=row_idx, column=3).value  # Counted (or Expected)
            col_d = ws.cell(row=row_idx, column=4).value  # Difference
            
            # Get category name
            if col_a and isinstance(col_a, str) and col_a.strip() and not col_a.startswith('='):
                category_name = col_a.strip()
                
                # Skip if not a valid category (filter out formula rows)
                if category_name in ['کۆمپیوتەر', 'مۆبیلیات', 'بارەگا', 'بینا']:
                    # Get or create category
                    if category_name not in categories_data:
                        category = Category.query.filter_by(name=category_name).first()
                        if not category:
                            category = Category(
                                name=category_name,
                                name_en={
                                    'کۆمپیوتەر': 'Computers',
                                    'مۆبیلیات': 'Vehicles',
                                    'بارەگا': 'Warehouses',
                                    'بینا': 'Buildings'
                                }.get(category_name, category_name)
                            )
                            db.session.add(category)
                            db.session.flush()
                        categories_data[category_name] = category
                    
                    # Create inventory item if we have data
                    if col_b and isinstance(col_b, (int, float)):
                        expected = int(col_b)
                        actual = int(col_c) if col_c and isinstance(col_c, (int, float)) else expected
                        
                        item = InventoryItem(
                            category_id=categories_data[category_name].id,
                            name=category_name,
                            system_name=f"{category_name}_{row_idx}",
                            expected_count=expected,
                            actual_count=actual,
                            notes=f"Imported from Excel - Discrepancy: {col_d}"
                        )
                        db.session.add(item)
        
        db.session.commit()
        print("✓ Data imported successfully!")
        print(f"✓ {len(categories_data)} categories created")

def init_demo_user():
    """Create a demo user for testing"""
    
    app = create_app()
    
    with app.app_context():
        # Check if user already exists
        if User.query.filter_by(username='admin').first():
            print("Admin user already exists")
            return
        
        # Create admin user
        admin = User(username='admin', is_admin=True)
        admin.set_password('admin123')
        db.session.add(admin)
        
        # Create demo user
        demo = User(username='demo', is_admin=False)
        demo.set_password('demo123')
        db.session.add(demo)
        
        db.session.commit()
        print("✓ Demo users created:")
        print("  - admin / admin123 (Admin)")
        print("  - demo / demo123 (Regular User)")

if __name__ == '__main__':
    excel_file = r'c:\Users\KDP\Desktop\inv-report-claude\ڕاپۆرتی_جەرد_v5.xlsx'
    
    print("Initializing inventory app...")
    print("📝 Creating demo users...")
    init_demo_user()
    
    print("📊 Importing Excel data...")
    if Path(excel_file).exists():
        import_from_excel(excel_file)
    else:
        print(f"Warning: Excel file not found at {excel_file}")
    
    print("\n✓ Setup complete!")
    print("Run: cd inventory-app && python run.py")
