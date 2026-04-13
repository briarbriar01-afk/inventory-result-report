from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash
from app import db
from app.models import User, Category, InventoryItem, AuditLog
from datetime import datetime

main_bp = Blueprint('main', __name__)
auth_bp = Blueprint('auth', __name__)
api_bp = Blueprint('api', __name__, url_prefix='/api')

# ==================== AUTHENTICATION ====================

@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('main.dashboard'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))

@auth_bp.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Passwords do not match', 'danger')
            return redirect(url_for('auth.register'))
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'danger')
            return redirect(url_for('auth.register'))
        
        user = User(username=username)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('auth.login'))
    
    return render_template('register.html')

# ==================== MAIN ROUTES ====================

@main_bp.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.login'))

@main_bp.route('/dashboard')
@login_required
def dashboard():
    categories = Category.query.all()
    total_expected = db.session.query(db.func.sum(InventoryItem.expected_count)).scalar() or 0
    total_actual = db.session.query(db.func.sum(InventoryItem.actual_count)).scalar() or 0
    total_discrepancy = total_actual - total_expected
    
    stats = {
        'total_items': InventoryItem.query.count(),
        'total_categories': Category.query.count(),
        'total_expected': total_expected,
        'total_actual': total_actual,
        'total_discrepancy': total_discrepancy,
    }
    
    return render_template('dashboard.html', categories=categories, stats=stats)

@main_bp.route('/inventory')
@login_required
def inventory():
    items = InventoryItem.query.all()
    categories = Category.query.all()
    return render_template('inventory.html', items=items, categories=categories)

@main_bp.route('/reports')
@login_required
def reports():
    categories = Category.query.all()
    
    report_data = []
    for category in categories:
        items = InventoryItem.query.filter_by(category_id=category.id).all()
        expected = sum([item.expected_count for item in items])
        actual = sum([item.actual_count for item in items])
        discrepancy = actual - expected
        
        report_data.append({
            'category': category.name,
            'expected': expected,
            'actual': actual,
            'discrepancy': discrepancy,
            'item_count': len(items),
        })
    
    return render_template('reports.html', report_data=report_data)

# ==================== API ROUTES ====================

@api_bp.route('/items', methods=['GET'])
@login_required
def get_items():
    category_id = request.args.get('category_id', type=int)
    
    query = InventoryItem.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    items = query.all()
    
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'category': item.category.name,
        'expected_count': item.expected_count,
        'actual_count': item.actual_count,
        'discrepancy': item.discrepancy,
        'notes': item.notes,
    } for item in items])

@api_bp.route('/items', methods=['POST'])
@login_required
def create_item():
    data = request.get_json()
    
    item = InventoryItem(
        category_id=data.get('category_id'),
        name=data.get('name'),
        system_name=data.get('system_name'),
        expected_count=data.get('expected_count', 0),
        actual_count=data.get('actual_count', 0),
        notes=data.get('notes'),
    )
    
    db.session.add(item)
    db.session.commit()
    
    log_audit(current_user.id, 'CREATE', item.id, None, f'{item.name}')
    
    return jsonify({'success': True, 'id': item.id}), 201

@api_bp.route('/items/<int:item_id>', methods=['PUT'])
@login_required
def update_item(item_id):
    item = InventoryItem.query.get_or_404(item_id)
    data = request.get_json()
    
    old_expected = item.expected_count
    old_actual = item.actual_count
    
    if 'name' in data:
        item.name = data['name']
    if 'expected_count' in data:
        item.expected_count = data['expected_count']
    if 'actual_count' in data:
        item.actual_count = data['actual_count']
    if 'notes' in data:
        item.notes = data['notes']
    
    db.session.commit()
    
    log_audit(current_user.id, 'UPDATE', item_id, 
              f'expected:{old_expected}, actual:{old_actual}',
              f'expected:{item.expected_count}, actual:{item.actual_count}')
    
    return jsonify({'success': True})

@api_bp.route('/items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_item(item_id):
    item = InventoryItem.query.get_or_404(item_id)
    
    log_audit(current_user.id, 'DELETE', item_id, item.name, None)
    
    db.session.delete(item)
    db.session.commit()
    
    return jsonify({'success': True})

@api_bp.route('/categories', methods=['GET'])
@login_required
def get_categories():
    categories = Category.query.all()
    return jsonify([{
        'id': cat.id,
        'name': cat.name,
        'name_en': cat.name_en,
    } for cat in categories])

@api_bp.route('/export/excel', methods=['GET'])
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'
    
    # Headers
    headers = ['Category', 'Item Name', 'Expected', 'Actual', 'Discrepancy', 'Notes']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    items = InventoryItem.query.all()
    for item in items:
        ws.append([
            item.category.name,
            item.name,
            item.expected_count,
            item.actual_count,
            item.discrepancy,
            item.notes or '',
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'inventory_{datetime.now().strftime("%Y%m%d")}.xlsx')

@api_bp.route('/reports/data', methods=['GET'])
@login_required
def get_reports_data():
    categories = Category.query.all()
    
    data = {
        'labels': [],
        'expected': [],
        'actual': [],
        'discrepancy': [],
    }
    
    for category in categories:
        items = InventoryItem.query.filter_by(category_id=category.id).all()
        expected = sum([item.expected_count for item in items])
        actual = sum([item.actual_count for item in items])
        
        data['labels'].append(category.name)
        data['expected'].append(expected)
        data['actual'].append(actual)
        data['discrepancy'].append(actual - expected)
    
    return jsonify(data)

def log_audit(user_id, action, item_id, old_value, new_value):
    log = AuditLog(
        user_id=user_id,
        action=action,
        item_id=item_id,
        old_value=old_value,
        new_value=new_value,
    )
    db.session.add(log)
    db.session.commit()
