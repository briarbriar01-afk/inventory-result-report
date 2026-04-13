#!/usr/bin/env python3
"""Initialize app data from Excel"""

import json
import openpyxl
from pathlib import Path
from werkzeug.security import generate_password_hash

# Data directory
DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)

# Initialize users
users = {
    '1': {
        'id': 1,
        'username': 'superadmin',
        'password_hash': generate_password_hash('superadmin123'),
        'is_admin': True,
        'role': 'superadmin'
    },
    '2': {
        'id': 2,
        'username': 'admin',
        'password_hash': generate_password_hash('admin123'),
        'is_admin': True,
        'role': 'admin'
    },
    '3': {
        'id': 3,
        'username': 'employee',
        'password_hash': generate_password_hash('employee123'),
        'is_admin': False,
        'role': 'employee'
    }
}

# Initialize categories
categories = {
    '1': {'id': 1, 'name': 'کۆمپیوتەر', 'name_en': 'Computers'},
    '2': {'id': 2, 'name': 'مۆبیلیات', 'name_en': 'Vehicles'},
    '3': {'id': 3, 'name': 'بارەگا', 'name_en': 'Warehouses'},
    '4': {'id': 4, 'name': 'بینا', 'name_en': 'Buildings'},
}

# Initialize items from Excel
items = {}
item_id = 1

try:
    excel_file = Path(__file__).parent.parent / 'ڕاپۆرتی_جەرد_v5.xlsx'
    if excel_file.exists():
        wb = openpyxl.load_workbook(str(excel_file))
        ws = wb['چارتەکان']
        
        category_map = {
            'کۆمپیوتەر': 1,
            'مۆبیلیات': 2,
            'بارەگا': 3,
            'بینا': 4,
        }
        
        for row_idx in range(2, ws.max_row + 1):
            category_name = ws.cell(row=row_idx, column=1).value
            cols_b = ws.cell(row=row_idx, column=2).value
            col_c = ws.cell(row=row_idx, column=3).value
            col_d = ws.cell(row=row_idx, column=4).value
            
            if category_name and isinstance(category_name, str) and category_name.strip() in category_map:
                if isinstance(cols_b, (int, float)):
                    expected = int(cols_b)
                    actual = int(col_c) if col_c and isinstance(col_c, (int, float)) else expected
                    
                    items[str(item_id)] = {
                        'id': str(item_id),
                        'category_id': category_map[category_name.strip()],
                        'name': category_name.strip(),
                        'expected_count': expected,
                        'actual_count': actual,
                        'notes': f'Imported - {col_d or ""}',
                        'created_at': '2026-04-13T00:00:00'
                    }
                    item_id += 1
except Exception as e:
    print(f"Note: Could not import from Excel: {e}")

# Save data
with open(DATA_DIR / 'users.json', 'w', encoding='utf-8') as f:
    json.dump(users, f, ensure_ascii=False, indent=2)

with open(DATA_DIR / 'categories.json', 'w', encoding='utf-8') as f:
    json.dump(categories, f, ensure_ascii=False, indent=2)

with open(DATA_DIR / 'items.json', 'w', encoding='utf-8') as f:
    json.dump(items, f, ensure_ascii=False, indent=2)

# Initialize reports (empty)
with open(DATA_DIR / 'reports.json', 'w', encoding='utf-8') as f:
    json.dump({}, f, ensure_ascii=False, indent=2)

print("✓ App initialized!")
print("\nDemo Users / بەکارهێنەرانی نمونە:")
print("  SuperAdmin: superadmin / superadmin123")
print("  Admin:      admin / admin123")
print("  Employee:   employee / employee123")
print(f"\n✓ Created {len(items)} inventory items from Excel")
