#!/usr/bin/env python3
"""
Simple Inventory Management App - Flask without SQLAlchemy
Uses JSON files for data storage
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from pathlib import Path
import json
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Data storage paths
DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)
USERS_FILE = DATA_DIR / 'users.json'
CATEGORIES_FILE = DATA_DIR / 'categories.json'
ITEMS_FILE = DATA_DIR / 'items.json'
REPORTS_FILE = DATA_DIR / 'reports.json'

# User model
class User(UserMixin):
    def __init__(self, id, username, password_hash=None, is_admin=False, role='employee'):
        self.id = id
        self.username = username
        self.password_hash = password_hash
        self.is_admin = is_admin
        self.role = role  # 'superadmin', 'admin', or 'employee'
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

@login_manager.user_loader
def load_user(user_id):
    users = load_users()
    for user_data in users.values():
        if user_data['id'] == int(user_id):
            return User(user_data['id'], user_data['username'], user_data['password_hash'], user_data.get('is_admin', False))
    return None

# Data management functions
def load_json(filepath):
    if filepath.exists():
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_json(filepath, data):
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_users():
    users = load_json(USERS_FILE)
    # Ensure all users have a role field
    for user_data in users.values():
        if 'role' not in user_data:
            user_data['role'] = 'employee'
    return users

def load_categories():
    return load_json(CATEGORIES_FILE)

def load_items():
    return load_json(ITEMS_FILE)

def save_users(users):
    save_json(USERS_FILE, users)

def save_categories(categories):
    save_json(CATEGORIES_FILE, categories)

def save_items(items):
    save_json(ITEMS_FILE, items)

def load_reports():
    return load_json(REPORTS_FILE)

def save_reports(reports):
    save_json(REPORTS_FILE, reports)

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        users = load_users()
        for user_data in users.values():
            if user_data['username'] == username:
                user = User(user_data['id'], user_data['username'], user_data['password_hash'], user_data.get('is_admin', False), user_data.get('role', 'employee'))
                if user.check_password(password):
                    login_user(user)
                    return redirect(url_for('dashboard'))
        
        flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Passwords do not match', 'danger')
            return redirect(url_for('register'))
        
        users = load_users()
        for user_data in users.values():
            if user_data['username'] == username:
                flash('Username already exists', 'danger')
                return redirect(url_for('register'))
        
        new_id = max([u['id'] for u in users.values()], default=0) + 1
        new_user = User(new_id, username)
        new_user.set_password(password)
        
        users[str(new_id)] = {
            'id': new_id,
            'username': username,
            'password_hash': new_user.password_hash,
            'is_admin': False
        }
        save_users(users)
        
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    categories = load_categories()
    items = load_items()
    
    stats = {
        'total_items': len(items),
        'total_categories': len(categories),
        'total_expected': sum(item['expected_count'] for item in items.values()),
        'total_actual': sum(item['actual_count'] for item in items.values()),
    }
    stats['total_discrepancy'] = stats['total_actual'] - stats['total_expected']
    
    return render_template('dashboard.html', categories=categories, items=items, stats=stats)

@app.route('/inventory')
@login_required
def inventory():
    categories = load_categories()
    items = load_items()
    return render_template('inventory.html', categories=categories, items=items)

@app.route('/reports')
@login_required
def reports():
    categories = load_categories()
    items = load_items()
    
    report_data = []
    for cat_id, category in categories.items():
        cat_items = [item for item in items.values() if item.get('category_id') == cat_id]
        expected = sum([item['expected_count'] for item in cat_items])
        actual = sum([item['actual_count'] for item in cat_items])
        
        report_data.append({
            'category': category['name'],
            'expected': expected,
            'actual': actual,
            'discrepancy': actual - expected,
            'item_count': len(cat_items),
        })
    
    return render_template('reports.html', report_data=report_data)

@app.route('/report-form')
@login_required
def report_form():
    """Generate official inventory report form"""
    return render_template('report_form.html')

# API routes
@app.route('/api/items', methods=['GET'])
@login_required
def get_items():
    items = load_items()
    return jsonify(list(items.values()))

@app.route('/api/items', methods=['POST'])
@login_required
def create_item():
    data = request.get_json()
    items = load_items()
    
    new_id = str(max([int(k) for k in items.keys()], default=0) + 1)
    items[new_id] = {
        'id': new_id,
        'category_id': data.get('category_id'),
        'name': data.get('name'),
        'expected_count': data.get('expected_count', 0),
        'actual_count': data.get('actual_count', 0),
        'notes': data.get('notes', ''),
        'created_at': datetime.now().isoformat()
    }
    save_items(items)
    
    return jsonify({'success': True, 'id': new_id}), 201

@app.route('/api/items/<item_id>', methods=['PUT'])
@login_required
def update_item(item_id):
    data = request.get_json()
    items = load_items()
    
    if item_id in items:
        items[item_id].update(data)
        items[item_id]['updated_at'] = datetime.now().isoformat()
        save_items(items)
        return jsonify({'success': True})
    
    return jsonify({'success': False}), 404

@app.route('/api/items/<item_id>', methods=['DELETE'])
@login_required
def delete_item(item_id):
    items = load_items()
    if item_id in items:
        del items[item_id]
        save_items(items)
        return jsonify({'success': True})
    
    return jsonify({'success': False}), 404

@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    categories = load_categories()
    return jsonify(list(categories.values()))

@app.route('/api/reports/data', methods=['GET'])
@login_required
def get_reports_data():
    categories = load_categories()
    items = load_items()
    
    data = {
        'labels': [],
        'expected': [],
        'actual': [],
        'discrepancy': [],
    }
    
    for cat_id, category in categories.items():
        cat_items = [item for item in items.values() if item.get('category_id') == cat_id]
        expected = sum([item['expected_count'] for item in cat_items])
        actual = sum([item['actual_count'] for item in cat_items])
        
        data['labels'].append(category['name'])
        data['expected'].append(expected)
        data['actual'].append(actual)
        data['discrepancy'].append(actual - expected)
    
    return jsonify(data)

@app.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    categories = load_categories()
    items = load_items()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'
    
    # Headers
    headers = ['Category', 'Item Name', 'Expected', 'Actual', 'Discrepancy', 'Notes']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for item in items.values():
        category_name = next((c['name'] for c in categories.values() if c['id'] == item.get('category_id')), 'Unknown')
        ws.append([
            category_name,
            item['name'],
            item['expected_count'],
            item['actual_count'],
            item['actual_count'] - item['expected_count'],
            item.get('notes', ''),
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'inventory_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/professional-report')
@login_required
def professional_report():
    """Generate professional inventory report form - Kurdish (Sorani)"""
    return render_template('new_report_form.html')

@app.route('/superadmin-dashboard')
@login_required
def superadmin_dashboard():
    """Superadmin dashboard to view all submitted reports"""
    user = current_user
    if user.role != 'superadmin':
        flash('Access denied. Superadmin only.', 'danger')
        return redirect(url_for('dashboard'))
    
    reports = load_reports()
    return render_template('superadmin_dashboard.html', reports=reports)

@app.route('/api/reports/save', methods=['POST'])
@login_required
def save_report():
    """Save report draft"""
    data = request.get_json()
    reports = load_reports()
    
    report_id = str(int(max([int(k) for k in reports.keys()], default=0)) + 1)
    
    reports[report_id] = {
        'id': report_id,
        'submitted_by': current_user.username,
        'user_id': current_user.id,
        'status': data.get('status', 'draft'),
        'created_at': datetime.now().isoformat(),
        'data': data
    }
    
    save_reports(reports)
    return jsonify({'success': True, 'id': report_id}), 201

@app.route('/api/reports/submit', methods=['POST'])
@login_required
def submit_report():
    """Submit report to superadmin"""
    data = request.get_json()
    reports = load_reports()
    
    report_id = str(int(max([int(k) for k in reports.keys()], default=0)) + 1)
    
    reports[report_id] = {
        'id': report_id,
        'submitted_by': current_user.username,
        'user_id': current_user.id,
        'status': 'submitted',
        'created_at': datetime.now().isoformat(),
        'submitted_at': datetime.now().isoformat(),
        'data': data
    }
    
    save_reports(reports)
    return jsonify({'success': True, 'id': report_id}), 201

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
